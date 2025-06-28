# app.py
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import os
import requests
import json
import pandas as pd
from datetime import datetime, timedelta
import io # For in-memory file operations
import sys # Import sys to check for PyInstaller frozen state
import tempfile # Added for creating temporary files for service account key
from openpyxl.styles import Font, PatternFill, Alignment # Import for Excel styling
import csv # Import for CSV operations

# Import Google Sheets libraries
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# --- Flask App Initialization ---
# Flask app instance will be created by create_app()
# app.secret_key and other configs will be set inside create_app()
# as recommended for PyInstaller compatibility.

# --- Configuration ---
# Admin credentials from environment variables (recommended for web deployment)
# These will be pulled from Render's environment variables
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin') # Default for local dev if .env missing
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'password123') # Default for local dev if .env missing

# Arkesel SMS API Key (from environment variable)
ARKESEL_API_KEY = os.environ.get('ARKESEL_API_KEY') # No default, MUST be set in env for SMS to work
ARKESEL_SENDER_ID = os.environ.get('ARKESEL_SENDER_ID', 'FarmApp') # Your registered sender ID

# Google Sheet Configuration - Use environment variable for the ID
# UPDATED GOOGLE_SHEET_ID to the latest one provided by you: 18NjH0VhNolUA3m_2JGvqR9oubcON92OVMQBxdf3Axi8
GOOGLE_SHEET_ID = os.environ.get('GOOGLE_SHEET_ID', "18NjH0VhNolUA3m_2JGvqR9oubcON92OVMQBxdf3Axi8")
# GOOGLE_SHEET_KEY_FILE is no longer used directly as key is reconstructed from env vars

# CSV Fallback Configuration
# Set to 'true' (case-insensitive) in environment variables to enable CSV fallback
USE_CSV_FALLBACK = os.environ.get('USE_CSV_FALLBACK', 'false').lower() == 'true'
CSV_FILE_NAME = 'farm_records.csv'
# Path for the CSV file. When bundled, sys._MEIPASS is the temp extraction dir.
# When deployed on Render, root_path is typically the /app directory.
# For local dev, it's the current directory.
# This will be set more precisely in create_app().
CSV_FILE_PATH = None 


# --- Google Sheets Integration ---
def init_google_sheets_client():
    """
    Initializes Google Sheets client by reconstructing the service account key
    from individual environment variables. This is suitable for cloud deployments.
    """
    print("--- DEBUG: init_google_sheets_client: Attempting to initialize Google Sheets client...")

    service_account_info = {}
    env_vars_to_check = [
        "GOOGLE_TYPE", "GOOGLE_PROJECT_ID", "GOOGLE_PRIVATE_KEY_ID", 
        "GOOGLE_PRIVATE_KEY", "GOOGLE_CLIENT_EMAIL", "GOOGLE_CLIENT_ID",
        "GOOGLE_AUTH_URI", "GOOGLE_TOKEN_URI", "GOOGLE_AUTH_PROVIDER_X509_CERT_URL",
        "GOOGLE_CLIENT_X509_CERT_URL", "GOOGLE_UNIVERSE_DOMAIN"
    ]

    for key in env_vars_to_check:
        value = os.environ.get(key)
        # Convert environment variable names (e.g., GOOGLE_PRIVATE_KEY) to JSON key names (e.g., private_key)
        service_account_info[key.lower().replace('google_', '')] = value
        # print(f"--- DEBUG: init_google_sheets_client: Env Var {key}: {value!r}") # Too verbose/sensitive

    # Special handling for private_key: replace escaped newlines (\\n) with actual newlines (\n)
    if service_account_info.get("private_key"):
        service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
        # Print only a snippet of the private key for security in logs
        # print(f"--- DEBUG: init_google_sheets_client: Private key after newline replacement (first 50 chars): {service_account_info['private_key'][:50]!r}...")

    # Validate critical parts - ensure private_key and client_email are present
    if not service_account_info.get("private_key") or not service_account_info.get("client_email"):
        print("--- DEBUG: init_google_sheets_client: ERROR: Missing critical Google service account environment variables (private_key or client_email are empty/None).")
        return None

    temp_key_file_path = None
    try:
        # Create a temporary JSON file from the environment variable data
        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json') as temp_key_file_obj:
            json.dump(service_account_info, temp_key_file_obj, indent=2)
            temp_key_file_path = temp_key_file_obj.name
        
        print(f"--- DEBUG: init_google_sheets_client: Temporary key file created at: {temp_key_file_path}")

        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_name(temp_key_file_path, scope)
        client = gspread.authorize(creds)
        print("--- DEBUG: init_google_sheets_client: Google Sheets client initialized successfully.")
        return client
    except FileNotFoundError:
        print(f"--- DEBUG: init_google_sheets_client: ERROR: Temporary service account key file not found at {temp_key_file_path}. This should not happen if created correctly.")
        return None
    except Exception as e:
        print(f"--- DEBUG: init_google_sheets_client: ERROR: Failed to initialize Google Sheets client: {e}")
        return None
    finally:
        # Clean up the temporary file (important for security)
        if temp_key_file_path and os.path.exists(temp_key_file_path):
            os.remove(temp_key_file_path)
            print(f"--- DEBUG: init_google_sheets_client: Removed temporary key file: {temp_key_file_path}")


def get_sheet(client, sheet_id):
    """Gets a specific worksheet using the spreadsheet ID."""
    try:
        # Open the spreadsheet by ID
        spreadsheet = client.open_by_key(sheet_id)
        # Get the first worksheet (default)
        worksheet = spreadsheet.sheet1
        print(f"--- DEBUG: get_sheet: Successfully opened sheet with ID: {sheet_id}")
        return worksheet
    except gspread.exceptions.SpreadsheetNotFound:
        print(f"--- DEBUG: get_sheet: ERROR: Spreadsheet with ID '{sheet_id}' not found. Please ensure the ID is correct and the sheet is shared with the service account.")
        return None
    except Exception as e:
        print(f"--- DEBUG: get_sheet: ERROR: Failed to open sheet with ID {sheet_id}: {e}")
        return None

def append_to_sheet(sheet, data):
    """Appends a row of data to the Google Sheet."""
    try:
        sheet.append_row(data)
        print(f"--- DEBUG: append_to_sheet: Successfully appended data to Google Sheet: {data}")
        return True
    except Exception as e:
        print(f"--- DEBUG: append_to_sheet: ERROR: Error appending data to sheet: {e}")
        return False

# --- CSV Helper Functions ---
def read_records_from_csv(file_path):
    """Reads all records from a CSV file and returns as a list of dictionaries."""
    print(f"--- DEBUG: read_records_from_csv: Attempting to read from CSV: {file_path}")
    records = []
    if not os.path.exists(file_path):
        print(f"--- DEBUG: read_records_from_csv: CSV file not found at {file_path}.")
        return records
    try:
        with open(file_path, mode='r', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                records.append(row)
        print(f"--- DEBUG: read_records_from_csv: Successfully read {len(records)} records from CSV.")
    except Exception as e:
        print(f"--- DEBUG: read_records_from_csv: ERROR reading CSV file {file_path}: {e}")
    return records

def write_records_to_csv(file_path, records_df):
    """Writes a pandas DataFrame to a CSV file."""
    print(f"--- DEBUG: write_records_to_csv: Attempting to write to CSV: {file_path}")
    try:
        # Ensure the directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        records_df.to_csv(file_path, index=False, encoding='utf-8')
        print(f"--- DEBUG: write_records_to_csv: Successfully wrote {len(records_df)} records to CSV.")
        return True
    except Exception as e:
        print(f"--- DEBUG: write_records_to_csv: ERROR writing CSV file {file_path}: {e}")
        return False

# --- Helper Functions for Data (Interacts with Google Sheets and CSV) ---
def save_record(record_type, data):
    """Saves a record to the Google Sheet and optionally to CSV."""
    
    # Attempt to save to Google Sheet first
    google_sheet_success = False
    client = init_google_sheets_client()
    if client:
        sheet = get_sheet(client, GOOGLE_SHEET_ID)
        if sheet:
            # Define the order of columns as expected in the Google Sheet
            ordered_headers = ['Date', 'Type', 'Category', 'Item', 'Quantity', 'Unit', 'Amount', 'Profit Per Unit', 'Total Profit']
            row_data = [data.get(header.replace(' ', '_').lower(), '') for header in ordered_headers]

            google_sheet_success = append_to_sheet(sheet, row_data)
            if not google_sheet_success:
                flash("Failed to add record to Google Sheet. Check server logs.", "danger")
        else:
            flash(f"Could not open Google Sheet with ID '{GOOGLE_SHEET_ID}'. Check server logs.", "danger")
    else:
        flash("Google Sheets client could not be initialized. Check server logs.", "danger")
    
    # If Google Sheet failed AND CSV fallback is enabled, or if CSV fallback is just enabled, save to CSV
    if not google_sheet_success and USE_CSV_FALLBACK:
        flash("Google Sheet save failed. Attempting to save to local CSV (data may not persist).", "warning")
    
    csv_success = False
    if USE_CSV_FALLBACK:
        # Prepare data for CSV
        df = pd.DataFrame([data])
        # Add missing columns with empty strings to ensure consistent CSV schema
        csv_columns = ['date', 'type', 'category', 'item', 'quantity', 'unit', 'amount', 'profit_per_unit', 'total_profit']
        for col in csv_columns:
            if col not in df.columns:
                df[col] = ''
        df = df[csv_columns] # Reorder columns for consistency

        # Read existing CSV, append new data, and write back
        existing_records_df = pd.DataFrame(read_records_from_csv(CSV_FILE_PATH))
        
        # Check if existing_records_df is empty and align columns for concatenation
        if existing_records_df.empty:
            final_df = df
        else:
            # Align columns before concat, fill missing with empty strings
            missing_in_existing = set(df.columns) - set(existing_records_df.columns)
            for col in missing_in_existing:
                existing_records_df[col] = ''
            
            missing_in_new = set(existing_records_df.columns) - set(df.columns)
            for col in missing_in_new:
                df[col] = ''
            
            # Ensure order before concatenation
            existing_records_df = existing_records_df[list(df.columns)]
            final_df = pd.concat([existing_records_df, df], ignore_index=True)

        csv_success = write_records_to_csv(CSV_FILE_PATH, final_df)
        if csv_success:
            flash("Record also saved to local CSV.", "info")
        else:
            flash("Failed to save record to local CSV.", "danger")

    return google_sheet_success or csv_success # Return true if either save method succeeded


def get_farm_statistics():
    """Retrieves aggregated farm data for dashboard statistics from Google Sheets or CSV fallback."""
    df = get_all_farm_records_df() # Use the unified data retrieval function
    if df.empty:
        print("--- DEBUG: get_farm_statistics: No records found for statistics.")
        return {
            'total_feeds_kg': 0,
            'total_expenditure': 0,
            'total_profit': 0,
            'layers_eggs_sold_crates': 0,
            'broilers_birds_sold': 0,
            'goats_sold': 0,
            'sheep_sold': 0
        }

    # Convert numeric columns to numeric, coercing errors
    for col in ['Quantity', 'Amount', 'Profit Per Unit', 'Total Profit']: # Use actual column names
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Rename columns to be lowercase for easier access in Python (optional, but good practice)
    # This might already be handled by get_all_farm_records_df
    df.columns = [col.replace(' ', '_').lower() for col in df.columns]

    stats = {
        'total_feeds_kg': df[df['type'] == 'feed_input']['quantity'].sum() if 'type' in df.columns and 'quantity' in df.columns else 0,
        'total_expenditure': df[df['type'] == 'expenditure']['amount'].sum() if 'type' in df.columns and 'amount' in df.columns else 0,
        'total_profit': df[df['type'] == 'profit']['total_profit'].sum() if 'type' in df.columns and 'total_profit' in df.columns else 0,
        'layers_eggs_sold_crates': df[(df['type'] == 'profit') & (df['category'] == 'Layers') & (df['item'] == 'Eggs Sold')]['quantity'].sum() if 'type' in df.columns and 'category' in df.columns and 'item' in df.columns and 'quantity' in df.columns else 0,
        'broilers_birds_sold': df[(df['type'] == 'profit') & (df['category'] == 'Broilers') & (df['item'] == 'Birds Sold')]['quantity'].sum() if 'type' in df.columns and 'category' in df.columns and 'item' in df.columns and 'quantity' in df.columns else 0,
        'goats_sold': df[(df['type'] == 'profit') & (df['category'] == 'Goats') & (df['item'] == 'Goat Meat')]['quantity'].sum() if 'type' in df.columns and 'category' in df.columns and 'item' in df.columns and 'quantity' in df.columns else 0,
        'sheep_sold': df[(df['type'] == 'profit') & (df['category'] == 'Sheep') & (df['item'] == 'Sheep Meat')]['quantity'].sum() if 'type' in df.columns and 'category' in df.columns and 'item' in df.columns and 'quantity' in df.columns else 0,
    }
    return stats

def get_all_farm_records_df():
    """Retrieves all farm records as a pandas DataFrame, prioritizing Google Sheets, with CSV as fallback."""
    print("--- DEBUG: get_all_farm_records_df: Called to retrieve all farm records.")
    
    records = []
    from_google_sheets = False

    # Attempt to retrieve from Google Sheets first
    client = init_google_sheets_client()
    if client:
        sheet = get_sheet(client, GOOGLE_SHEET_ID)
        if sheet:
            try:
                records = sheet.get_all_records()
                if records:
                    print(f"--- DEBUG: get_all_farm_records_df: Successfully retrieved {len(records)} records from Google Sheet.")
                    from_google_sheets = True
                else:
                    print("--- DEBUG: get_all_farm_records_df: Google Sheet is empty.")
                    flash("No records found in the Google Sheet.", "info")
            except Exception as e:
                print(f"--- DEBUG: get_all_farm_records_df: ERROR retrieving records from Google Sheet: {e}")
                flash("Error retrieving records from Google Sheet. Check server logs.", "danger")
        else:
            flash(f"Could not open Google Sheet with ID '{GOOGLE_SHEET_ID}'. Check server logs.", "danger")
    else:
        flash("Google Sheets client could not be initialized for record retrieval.", "danger")

    # If Google Sheets failed or is empty and CSV fallback is enabled, try CSV
    if (not from_google_sheets or not records) and USE_CSV_FALLBACK:
        flash("Falling back to local CSV for records retrieval. Data may not be up-to-date or persist.", "warning")
        print("--- DEBUG: get_all_farm_records_df: Attempting to read records from CSV fallback.")
        records = read_records_from_csv(CSV_FILE_PATH)
        if not records:
            flash("No records found in local CSV.", "info")

    if not records:
        print("--- DEBUG: get_all_farm_records_df: No records found from any source, returning empty DataFrame.")
        return pd.DataFrame()

    df = pd.DataFrame(records)
    print(f"--- DEBUG: Initial DataFrame shape: {df.shape}")
    print(f"--- DEBUG: Initial DataFrame columns (raw from source): {df.columns.tolist()}")
    print(f"--- DEBUG: Initial DataFrame head:\n{df.head().to_string()}")

    # Define desired standardized column names and common variations (lowercase, no spaces)
    standardized_column_map = {
        'Date': ['date', 'recorddate', 'transactiondate', 'timestamp'],
        'Type': ['type', 'recordtype', 'recordkind', 'transactiontype'],
        'Category': ['category', 'itemcategory', 'classification'],
        'Item': ['item', 'description', 'product', 'detail'],
        'Quantity': ['quantity', 'qty', 'amountbought', 'amountsold'],
        'Unit': ['unit', 'uom', 'measure'],
        'Amount': ['amount', 'expenditureamount', 'cost', 'totalcost', 'value'], # For expenditure
        'Profit Per Unit': ['profitperunit', 'ppu', 'unitprofit', 'priceperunit'],
        'Total Profit': ['totalprofit', 'profit', 'netsales', 'revenue']
    }

    # Normalize existing column names to facilitate matching
    normalized_df_columns = {col.lower().replace(' ', ''): col for col in df.columns}
    
    # Create a mapping from current actual column names to desired standardized names
    column_renaming_dict = {}
    final_columns = set() # To keep track of columns we actually want to keep/rename to

    for desired_name, variations in standardized_column_map.items():
        found_match = False
        for var in variations:
            if var in normalized_df_columns:
                original_matched_col = normalized_df_columns[var]
                if original_matched_col != desired_name:
                    column_renaming_dict[original_matched_col] = desired_name
                else: # It's already the desired_name (case-sensitive)
                    column_renaming_dict[original_matched_col] = desired_name # Ensure it's in the dict
                found_match = True
                final_columns.add(desired_name)
                break
        
        # If a desired column is not found through variations or exact match
        if not found_match and desired_name not in df.columns:
            df[desired_name] = '' # Add as empty to prevent KeyError
            final_columns.add(desired_name) # Add to final columns list
            print(f"--- DEBUG: Critical column '{desired_name}' not found, added as empty.")


    # Apply the renaming
    if column_renaming_dict:
        df.rename(columns=column_renaming_dict, inplace=True)
    
    # Reorder columns to a consistent display order (optional but good for consistency)
    ordered_display_columns = [
        'Date', 'Type', 'Category', 'Item', 'Quantity', 'Unit', 
        'Amount', 'Profit Per Unit', 'Total Profit'
    ]
    
    # Filter for columns that actually exist in the DataFrame (including newly added empty ones)
    df_columns_as_set = set(df.columns)
    columns_to_reorder = [col for col in ordered_display_columns if col in df_columns_as_set]
    
    # Add any original columns that weren't part of our standardized map to the end
    for col in df.columns:
        if col not in columns_to_reorder:
            columns_to_reorder.append(col)

    df = df[columns_to_reorder]
    
    print(f"--- DEBUG: After standardization and adding missing, DataFrame columns: {df.columns.tolist()}")
    print(f"--- DEBUG: DataFrame head after column processing:\n{df.head().to_string()}")

    # Check for critical columns and flash warnings if they were added as empty
    critical_columns_for_warning = ['Date', 'Type', 'Amount', 'Total Profit']
    for col_name in critical_columns_for_warning:
        if col_name in df.columns and df[col_name].empty and not records: # Only flash if df is entirely empty from the start
            continue # Don't flash if no records at all
        
        # Check if the column exists and if it's uniformly empty/missing in a non-empty df
        if col_name in df.columns and not df.empty and df[col_name].astype(str).str.strip().eq('').all():
             flash_message = f"Warning: The '{col_name}' column appears to be empty or was not explicitly found in your data source. Report data for this column might be missing or inaccurate. Please ensure your Google Sheet/CSV has a column specifically named '{col_name}' or a common variation."
             if 'Date' in col_name:
                 flash_message = f"Warning: The '{col_name}' column appears to be empty or was not explicitly found. Please ensure your Google Sheet/CSV has a column specifically named '{col_name}' for accurate reporting."
             elif 'Type' in col_name:
                  flash_message = f"Warning: The '{col_name}' column appears to be empty or was not explicitly found. Report filtering might be inaccurate. Please ensure your Google Sheet/CSV has a column specifically named '{col_name}'."
             flash(flash_message, "warning")


    # Ensure 'Date' column is in datetime format AFTER ensuring it exists and is named correctly
    if 'Date' in df.columns:
        print(f"--- DEBUG: 'Date' column dtype (before convert): {df['Date'].dtype}")
        print(f"--- DEBUG: 'Date' column values (before convert, head):\n{df['Date'].head().to_string()}")

        initial_rows_before_dropna = df.shape[0]
        
        # Convert date column, coercing errors to NaT
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        
        print(f"--- DEBUG: After pd.to_datetime, 'Date' column dtype: {df['Date'].dtype}")
        print(f"--- DEBUG: 'Date' column values (after convert, head):\n{df['Date'].head().to_string()}")
        print(f"--- DEBUG: Count of NaT values in 'Date' column: {df['Date'].isna().sum()}")

        # Drop rows where 'Date' became NaT (invalid date)
        df.dropna(subset=['Date'], inplace=True)
        
        print(f"--- DEBUG: After dropna(subset=['Date']), DataFrame shape: {df.shape}")
        if df.shape[0] < initial_rows_before_dropna:
            dropped_rows_count = initial_rows_before_dropna - df.shape[0]
            print(f"--- DEBUG: {dropped_rows_count} rows dropped due to invalid 'Date' values.")
            if dropped_rows_count > 0 and df.empty:
                 flash(f"Warning: All records were removed because their 'Date' column contained invalid or empty date formats. Please check your Google Sheet/CSV.", "warning")
            elif dropped_rows_count > 0:
                 flash(f"Warning: Some records were removed because their 'Date' column contained invalid or empty date formats. Please check your Google Sheet/CSV. Remaining records: {df.shape[0]}", "warning")

    else:
        print("--- DEBUG: 'Date' column still missing or invalid after all checks, returning empty DataFrame.")
        flash("Error: Failed to establish a valid 'Date' column. Reports cannot be generated. Please ensure your Google Sheet/CSV has a column for dates (e.g., 'Date').", "danger")
        return pd.DataFrame()

    # Convert relevant numeric columns after date processing, as errors='coerce' might be needed
    # for columns that might have mixed types from Google Sheets.
    for col in ['Quantity', 'Amount', 'Profit Per Unit', 'Total Profit']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    print(f"--- DEBUG: Final DataFrame shape being returned: {df.shape}")
    print(f"--- DEBUG: Final DataFrame head being returned:\n{df.head().to_string()}")
    return df

def update_record_in_sheet(row_index_in_sheet, updated_data_dict):
    """Updates a specific row in the Google Sheet and optionally in local CSV."""
    google_sheet_success = False
    
    client = init_google_sheets_client()
    if client:
        sheet = get_sheet(client, GOOGLE_SHEET_ID)
        if sheet:
            try:
                # Define the exact order of columns as they appear in the Google Sheet headers.
                ordered_headers = ['Date', 'Type', 'Category', 'Item', 'Quantity', 'Unit', 'Amount', 'Profit Per Unit', 'Total Profit']
                row_values = [updated_data_dict.get(header.replace(' ', '_').lower(), '') for header in ordered_headers]
                
                # Update in Google Sheet
                range_name = f'A{row_index_in_sheet}:{chr(ord("A") + len(ordered_headers) - 1)}{row_index_in_sheet}'
                sheet.update(range_name, [row_values])
                print(f"--- DEBUG: update_record_in_sheet: Successfully updated row {row_index_in_sheet} in Google Sheet.")
                google_sheet_success = True
                flash('Record updated successfully in Google Sheet!', 'success')
            except Exception as e:
                print(f"--- DEBUG: update_record_in_sheet: ERROR updating Google Sheet row {row_index_in_sheet}: {e}")
                flash('Failed to update record in Google Sheet. Check server logs.', 'danger')
        else:
            flash(f"Could not open Google Sheet with ID '{GOOGLE_SHEET_ID}' for update. Check server logs.", "danger")
    else:
        flash("Google Sheets client could not be initialized for update. Check server logs.", "danger")

    csv_success = False
    if USE_CSV_FALLBACK:
        if not google_sheet_success: # Only flash this warning if Google Sheet update failed
            flash("Google Sheet update failed. Attempting to update local CSV (data may not persist).", "warning")
        
        all_records_df = pd.DataFrame(read_records_from_csv(CSV_FILE_PATH))
        
        # Google Sheet row_index_in_sheet is 1-indexed for headers + actual row
        # DataFrame index is 0-indexed.
        # So df_index = row_index_in_sheet - 2
        df_index = row_index_in_sheet - 2 

        if not all_records_df.empty and 0 <= df_index < len(all_records_df):
            # Update the DataFrame at the specific index
            for key, value in updated_data_dict.items():
                if key in all_records_df.columns.str.lower().str.replace('_', ' '):
                    col_name_in_df = all_records_df.columns[all_records_df.columns.str.lower().str.replace('_', ' ') == key].iloc[0]
                    all_records_df.at[df_index, col_name_in_df] = value
                
            csv_success = write_records_to_csv(CSV_FILE_PATH, all_records_df)
            if csv_success:
                flash("Record also updated in local CSV.", "info")
            else:
                flash("Failed to update record in local CSV.", "danger")
        elif not all_records_df.empty:
            print(f"--- DEBUG: update_record_in_sheet: DataFrame index {df_index} out of bounds for CSV update (df size: {len(all_records_df)})")
            flash("Could not find record in local CSV for update. It might not exist there yet.", "warning")
        else:
             print(f"--- DEBUG: update_record_in_sheet: No records in CSV for update.")
             flash("No records in local CSV to update.", "info")

    return google_sheet_success or csv_success # Return true if either update method succeeded


def create_app():
    """
    Creates and configures the Flask application instance.
    All routes and app-specific configurations should be defined here.
    """
    # Create the Flask app instance
    app_instance = Flask(__name__)
    print(f"--- DEBUG: app.py: Flask app instance created inside create_app(): {id(app_instance)}")

    # When bundled by PyInstaller, set the root_path to the temporary extraction directory
    # For web deployments (like Render), this will be the base directory of the app.
    if getattr(sys, 'frozen', False):
        app_instance.root_path = sys._MEIPASS
        print(f"--- DEBUG: app.py: Running in PyInstaller bundle. app_instance.root_path set to: {app_instance.root_path}")
    else:
        # In a typical Render/Gunicorn deployment, this will be the current working directory
        # where your app.py resides.
        app_instance.root_path = os.path.dirname(os.path.abspath(__file__))
        print(f"--- DEBUG: app.py: Running in normal environment. app_instance.root_path: {app_instance.root_path}")

    # Set CSV_FILE_PATH globally here after app_instance.root_path is determined
    global CSV_FILE_PATH
    CSV_FILE_PATH = os.path.join(app_instance.root_path, CSV_FILE_NAME)
    print(f"--- DEBUG: CSV_FILE_PATH set to: {CSV_FILE_PATH}")
    print(f"--- DEBUG: USE_CSV_FALLBACK is set to: {USE_CSV_FALLBACK}")


    # Flask Secret Key from environment variable (recommended for web deployment)
    app_instance.secret_key = os.environ.get('FLASK_SECRET_KEY', 'a_default_secret_key_if_not_set_in_env_for_dev_only')

    # --- Routes ---
    print("--- DEBUG: app.py: Registering routes inside create_app()...")

    @app_instance.route('/')
    def index():
        print("--- DEBUG: app.py: index() route called.")
        return render_template('index.html')

    @app_instance.route('/login', methods=['GET', 'POST'])
    def login():
        print("--- DEBUG: app.py: login() route called.")
        if request.method == 'POST':
            username = request.form['username']
            password = request.form['password']
            
            # --- START NEW DEBUG PRINTS ---
            print(f"--- DEBUG: Login Debug - Raw Received username: {username!r} (len: {len(username)})")
            print(f"--- DEBUG: Login Debug - Raw Received password: {password!r} (len: {len(password)})")
            print(f"--- DEBUG: Login Debug - Raw Expected username: {ADMIN_USERNAME!r} (len: {len(ADMIN_USERNAME)})")
            print(f"--- DEBUG: Login Debug - Raw Expected password: {ADMIN_PASSWORD!r} (len: {len(ADMIN_PASSWORD)})")
            print(f"--- DEBUG: Login Debug - Username comparison: {username == ADMIN_USERNAME}")
            print(f"--- DEBUG: Login Debug - Password comparison: {password == ADMIN_PASSWORD}")
            # --- END NEW DEBUG PRINTS ---

            if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
                session['logged_in'] = True
                flash('Logged in successfully!', 'success')
                print("--- DEBUG: Login successful.")
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Invalid credentials. Please try again.', 'danger')
                print("--- DEBUG: Login failed: Invalid credentials.")
        return render_template('login.html')

    @app_instance.route('/logout')
    def logout():
        print("--- DEBUG: app.py: logout() route called.")
        session.pop('logged_in', None)
        flash('You have been logged out.', 'info')
        return redirect(url_for('index'))

    @app_instance.before_request
    def require_login():
        print(f"--- DEBUG: app.py: before_request called for endpoint: {request.endpoint}")
        if request.endpoint in ['admin_dashboard', 'view_records', 'export_records', 'edit_record', 'view_monthly_report', 'view_weekly_report', 'add_record', 'send_custom_sms'] and not session.get('logged_in'):
            flash('Please log in to access this page.', 'warning')
            print(f"--- DEBUG: Redirecting to login for endpoint: {request.endpoint}")
            return redirect(url_for('login'))

    @app_instance.route('/admin')
    def admin_dashboard():
        print("--- DEBUG: app.py: admin_dashboard() route called.")
        stats = get_farm_statistics()
        return render_template('admin.html', stats=stats)

    @app_instance.route('/admin/add_record', methods=['POST'])
    def add_record():
        print("--- DEBUG: app.py: add_record() route called.")
        if not session.get('logged_in'):
            flash('Unauthorized access.', 'danger')
            return redirect(url_for('login'))

        record_type = request.form['record_type']
        data = {'date': datetime.now().strftime('%Y-%m-%d')}

        try:
            if record_type == 'feed':
                data['type'] = 'feed_input'
                data['category'] = request.form['feed_category']
                data['item'] = request.form['feed_type']
                data['quantity'] = float(request.form['feed_quantity'])
                data['unit'] = 'kg'
                success = save_record('feed', data)
                if success:
                    flash('Feed record added successfully!', 'success')
                else:
                    # Flash message already handled inside save_record
                    pass
            elif record_type == 'expenditure':
                data['type'] = 'expenditure'
                data['category'] = request.form['exp_category']
                data['item'] = request.form['exp_item']
                data['amount'] = float(request.form['exp_amount'])
                success = save_record('expenditure', data)
                if success:
                    flash('Expenditure record added successfully!', 'success')
                else:
                    # Flash message already handled inside save_record
                    pass
            elif record_type == 'profit':
                data['type'] = 'profit'
                data['category'] = request.form['profit_category']
                data['item'] = request.form['profit_item']
                data['quantity'] = int(request.form['profit_quantity'])
                data['profit_per_unit'] = float(request.form['profit_per_unit'])
                data['total_profit'] = data['quantity'] * data['profit_per_unit']
                data['unit'] = 'crates' if 'Eggs' in data['item'] else ('birds' if 'Birds' in data['item'] else 'units')
                success = save_record('profit', data)
                if success:
                    flash('Profit record added successfully!', 'success')
                else:
                    # Flash message already handled inside save_record
                    pass
            else:
                flash('Invalid record type.', 'danger')
        except ValueError:
            flash('Invalid input for quantity, amount, or profit per unit. Please enter numbers.', 'danger')
        except Exception as e:
            flash(f'An unexpected error occurred: {e}', 'danger')
            print(f"Error adding record: {e}")

        return redirect(url_for('admin_dashboard'))

    @app_instance.route('/admin/send_sms', methods=['POST'])
    def send_custom_sms():
        print("--- DEBUG: app.py: send_custom_sms() route called.")
        if not session.get('logged_in'):
            flash('Unauthorized access.', 'danger')
            return redirect(url_for('login'))

        recipient = request.form['recipient_number']
        message = request.form['sms_message']

        if not recipient or not message:
            flash('Recipient number and message are required!', 'warning')
            return redirect(url_for('admin_dashboard'))

        # Assuming send_sms function is defined elsewhere or will be added.
        # For now, this will just print a debug message.
        # In a real app, integrate Arkesel API call here.
        print(f"--- DEBUG: send_custom_sms: Attempting to send SMS to {recipient} with message: {message}")
        # Placeholder for actual SMS sending logic
        try:
            # Placeholder for actual SMS API call using ARKESEL_API_KEY and ARKESEL_SENDER_ID
            # Example:
            # response = requests.post(
            #     'https://sms.arkesel.com/sms/api',
            #     data={
            #         'action': 'send-sms',
            #         'api_key': ARKESEL_API_KEY,
            #         'to': recipient,
            #         'from': ARKESEL_SENDER_ID,
            #         'sms': message
            #     }
            # )
            # response.raise_for_status() # Raise an HTTPError for bad responses (4xx or 5xx)
            # result = response.json()
            # if result.get('status') == 'success': # Adjust based on Arkesel's actual success indicator
            #     flash(f'SMS sent successfully to {recipient}!', 'success')
            # else:
            #     flash(f"Failed to send SMS: {result.get('message', 'Unknown error')}", 'danger')
            flash(f'SMS functionality is a placeholder. SMS to {recipient} with message: "{message}"', 'info')
            success = True # Assume success for placeholder
            msg = "SMS simulated successfully."
        except requests.exceptions.RequestException as e:
            flash(f'Error connecting to SMS service: {e}', 'danger')
            success = False
            msg = f"Connection error: {e}"
        except Exception as e:
            flash(f'An unexpected error occurred while sending SMS: {e}', 'danger')
            success = False
            msg = f"Unexpected error: {e}"

        return redirect(url_for('admin_dashboard'))

    @app_instance.route('/admin/view_records')
    def view_records():
        print("--- DEBUG: app.py: view_records() route called.")
        if not session.get('logged_in'):
            flash('Please log in to view records.', 'warning')
            return redirect(url_for('login'))

        df_records = get_all_farm_records_df()
        if df_records.empty:
            flash("No records available to display.", "info")
            return render_template('view_records.html', records=[], columns=[])

        records_list = df_records.to_dict(orient='records')
        columns = df_records.columns.tolist()

        return render_template('view_records.html', records=records_list, columns=columns)

    @app_instance.route('/admin/edit_record/<int:record_index>', methods=['GET', 'POST'])
    def edit_record(record_index):
        print("--- DEBUG: app.py: edit_record() route called.")
        if not session.get('logged_in'):
            flash('Please log in to edit records.', 'warning')
            return redirect(url_for('login'))

        df_records = get_all_farm_records_df()
        if df_records.empty:
            flash("No records found to edit.", "danger")
            return redirect(url_for('view_records'))
        
        # Adjust index for DataFrame (0-indexed)
        # Assuming record_index from URL is 0-indexed based on rendered list
        if record_index < 0 or record_index >= len(df_records):
            flash("Record not found for editing.", "danger")
            return redirect(url_for('view_records'))

        record_to_edit = df_records.iloc[record_index].to_dict()
        # Ensure keys are lowercase and snake_case for consistency with form data
        formatted_record = {k.replace(' ', '_').lower(): v for k, v in record_to_edit.items()}
        
        # The sheet_row_number is 1-indexed and accounts for header row
        # If your records are from df.get_all_records() which skips headers,
        # then the sheet row for df.iloc[index] is index + 2
        sheet_row_number = record_index + 2 


        if request.method == 'POST':
            updated_data = {
                'date': request.form['date'],
                'type': request.form['type'],
                'category': request.form['category'],
                'item': request.form['item'],
                'quantity': request.form.get('quantity', ''),
                'unit': request.form.get('unit', ''),
                'amount': request.form.get('amount', ''),
                'profit_per_unit': request.form.get('profit_per_unit', ''),
                'total_profit': request.form.get('total_profit', '')
            }
            
            for key in ['quantity', 'amount', 'profit_per_unit', 'total_profit']:
                if updated_data[key]:
                    try:
                        updated_data[key] = float(updated_data[key])
                        if key == 'quantity': # Quantity can be integer in some cases
                            updated_data[key] = int(updated_data[key])
                    except ValueError:
                        flash(f"Invalid number for {key.replace('_', ' ').title()}. Please enter a valid number.", "danger")
                        return render_template('edit_record.html', record=formatted_record, record_index=record_index)
                else:
                    updated_data[key] = '' # Ensure empty string for missing/invalid numeric fields

            if updated_data.get('quantity') and updated_data.get('profit_per_unit'):
                try:
                    updated_data['total_profit'] = float(updated_data['quantity']) * float(updated_data['profit_per_unit'])
                except ValueError:
                    updated_data['total_profit'] = '' # Set to empty if calculation fails

            success = update_record_in_sheet(sheet_row_number, updated_data)
            if success:
                flash('Record updated successfully!', 'success')
                return redirect(url_for('view_records'))
            else:
                # Flash message already handled inside update_record_in_sheet
                pass
                return render_template('edit_record.html', record=formatted_record, record_index=record_index)

        return render_template('edit_record.html', record=formatted_record, record_index=record_index)


    @app_instance.route('/admin/export_records')
    def export_records():
        print("--- DEBUG: app.py: export_records() route called.")
        if not session.get('logged_in'):
            flash('Please log in to export records.', 'warning')
            return redirect(url_for('login'))

        df_records = get_all_farm_records_df()
        if df_records.empty:
            flash("No records available to export.", "warning")
            return redirect(url_for('view_records'))

        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Farm Records"

        headers = df_records.columns.tolist()
        sheet.append(headers)

        # Apply header styling
        # Font and PatternFill need to be imported from openpyxl.styles
        # Make sure these are imported at the top:
        # from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for r_idx, row in df_records.iterrows():
            row_data = row.tolist()
            sheet.append(row_data)

        for column in sheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_name].width = adjusted_width


        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Farm_Records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    @app_instance.route('/admin/reports/monthly')
    def view_monthly_report():
        print("--- DEBUG: app.py: view_monthly_report() route called.")
        if not session.get('logged_in'):
            flash('Please log in to view reports.', 'warning')
            return redirect(url_for('login'))

        df = get_all_farm_records_df()
        if df.empty:
            flash("No records available for reports.", "info")
            return render_template('monthly_report.html', report_data={'month': datetime.now().strftime('%B %Y'), 'total_profit': 0, 'total_expenditure': 0, 'records': []}, report_title="Monthly Report")

        # Ensure 'Date' column is in datetime format and valid
        if 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
            df.dropna(subset=['Date'], inplace=True)
        
        if df.empty: # Check again after dropping invalid dates
            flash("No valid date records available for this report after filtering invalid dates.", "warning")
            return render_template('monthly_report.html', report_data={'month': datetime.now().strftime('%B %Y'), 'total_profit': 0, 'total_expenditure': 0, 'records': []}, report_title="Monthly Report")


        current_month = datetime.now().month
        current_year = datetime.now().year
        monthly_records = df[(df['Date'].dt.month == current_month) & (df['Date'].dt.year == current_year)]

        # Ensure numeric columns are actually numeric before summing
        for col in ['Amount', 'Total Profit']:
            if col in monthly_records.columns:
                monthly_records[col] = pd.to_numeric(monthly_records[col], errors='coerce').fillna(0)


        total_monthly_profit = monthly_records[monthly_records['Type'] == 'Profit']['Total Profit'].sum()
        total_monthly_expenditure = monthly_records[monthly_records['Type'] == 'Expenditure']['Amount'].sum()

        report_data = {
            'month': datetime.now().strftime('%B %Y'),
            'total_profit': total_monthly_profit,
            'total_expenditure': total_monthly_expenditure,
            'records': monthly_records.to_dict(orient='records')
        }

        return render_template('monthly_report.html', report_data=report_data, report_title="Monthly Profit & Expenditure Report")


    @app_instance.route('/admin/reports/weekly')
    def view_weekly_report():
        print("--- DEBUG: app.py: view_weekly_report() route called.")
        if not session.get('logged_in'):
            flash('Please log in to view reports.', 'warning')
            return redirect(url_for('login'))

        df = get_all_farm_records_df()
        if df.empty:
            flash("No records available for reports.", "info")
            return render_template('weekly_report.html', report_data={'week_range': 'Current Week', 'total_profit': 0, 'total_expenditure': 0, 'records': []}, report_title="Weekly Report")

        # Ensure 'Date' column is in datetime format and valid
        if 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
            df.dropna(subset=['Date'], inplace=True)

        if df.empty: # Check again after dropping invalid dates
            flash("No valid date records available for this report after filtering invalid dates.", "warning")
            return render_template('weekly_report.html', report_data={'week_range': 'Current Week', 'total_profit': 0, 'total_expenditure': 0, 'records': []}, report_title="Weekly Report")


        today = datetime.now().date()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        weekly_records = df[(df['Date'].dt.date >= start_of_week) & (df['Date'].dt.date <= end_of_week)]

        # Ensure numeric columns are actually numeric before summing
        for col in ['Amount', 'Total Profit']:
            if col in weekly_records.columns:
                weekly_records[col] = pd.to_numeric(weekly_records[col], errors='coerce').fillna(0)

        total_weekly_profit = weekly_records[weekly_records['Type'] == 'Profit']['Total Profit'].sum()
        total_weekly_expenditure = weekly_records[weekly_records['Type'] == 'Expenditure']['Amount'].sum()

        report_data = {
            'week_range': f"{start_of_week.strftime('%Y-%m-%d')} to {end_of_week.strftime('%Y-%m-%d')}",
            'total_profit': total_weekly_profit,
            'total_expenditure': total_weekly_expenditure,
            'records': weekly_records.to_dict(orient='records')
        }

        return render_template('weekly_report.html', report_data=report_data, report_title="Weekly Profit & Expenditure Report")

    return app_instance

# This line is CRUCIAL. It calls create_app() and assigns the configured app
# instance to the global 'app' variable when app.py is imported.
app = create_app()

# This block ensures the Flask development server runs when app.py is executed directly.
if __name__ == '__main__':
    # When running locally, ensure .env is loaded
    from dotenv import load_dotenv
    load_dotenv() 
    app.run(debug=True, host='0.0.0.0', port=5000)
